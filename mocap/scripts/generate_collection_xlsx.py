"""Generate the hand-teleop mocap-vs-fastsam3d data-collection workbook.

    python3 mocap_validation/generate_collection_xlsx.py [out.xlsx]

Sheets: Protocol, Tasks, Subjects, Data log (pre-filled), Schedule.
Subject names are left blank on purpose (filled in by the operator).
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

N_SUBJECTS = 15
N_TAKES = 2          # boss: ≥2 takes per task, 10 reps each = ≥20 reps/task
HDR = PatternFill("solid", fgColor="2F5597")
HDR_FONT = Font(bold=True, color="FFFFFF")
SUB = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# id, name, description, occlusion, reps, dur, why
TASKS = [
    ("C0", "Calibration: T-pose + ROM",
     "Hand still & flat with ALL markers visible ~2 s (seed + neutral pose), "
     "then slow full flex/extend of each finger in turn (bone-length + STA "
     "reference). One take.",
     "None", 1, "~60 s", "Neutral reference (Point 3) + bone lengths + STA."),
    ("T0", "Hand flexion + per-finger pinch",
     "From open hand: flex all fingers to a fist and extend (full ROM); then "
     "pinch the thumb tip to each fingertip in turn — index, middle, ring, "
     "pinky. One rep = full flex/extend + the four pinches.",
     "None / low", 10, "~12 s/rep",
     "Accuracy ceiling (flexion) + per-finger opposition; teleop-critical."),
    ("T1", "Pick-and-place",
     "Grasp a ~4 cm cube/cylinder, move it between two marks ~20 cm apart, "
     "release. Alternate direction each rep.",
     "Medium", 10, "~6 s/rep",
     "Canonical manipulation: reach–grasp–transport–release."),
    ("T2", "Screwdriver",
     "Insert a screwdriver on a screw and turn ~3 half-turns with regrips.",
     "Medium-high + wrist rotation", 10, "~10 s/rep",
     "Tool use + pronation/supination; dynamic occlusion."),
    ("T3", "Open a plastic bottle cap",
     "Unscrew the cap of a plastic bottle; the other (unmarked) hand holds "
     "the bottle.",
     "High", 10, "~8 s/rep", "In-hand ADL; occlusion stress-test."),
    ("T4", "Hammer (slowly)",
     "SLOW, controlled hammer taps on a peg/nail — no fast swings (protects "
     "the markers; the informative signal is the power grip, not the speed).",
     "Medium (power)", 10, "~8 s/rep",
     "Power grasp; keep it slow so markers survive and stay tracked."),
]

# per-subject scheduling grid (days × time slots)
CAL_DAYS = ["Wed 23 Jul", "Thu 24 Jul", "Fri 25 Jul"]
CAL_SLOTS = ["09:00–10:30", "10:30–12:00", "12:00–13:00 (lunch)",
             "13:00–14:30", "14:30–16:00", "16:00–17:30"]

SUBJ_COLS = ["Subject ID", "Name (fill in)", "Date", "Session start",
             "Session end", "Handedness", "Age", "Sex",
             "Hand length mm\n(wrist crease→middle tip)",
             "Index length mm", "Bracelet/rings removed?", "Notes"]

LOG_COLS = ["Subject ID", "Take type", "Task ID", "Take #", "Mocap file (.c3d)",
            "RGB/cosmik file", "Sync clap Y/N", "# reps", "Marker dropout Y/N",
            "Marker fell off Y/N", "Quality", "Redo? Y/N", "Notes"]

SCHEDULE = [
    ("0:00", "0:10", "Consent + anthropometry (hand/finger lengths, handedness)"),
    ("0:10", "0:30", "Marker placement: 21 dorsal hand markers + wrist; secure "
                     "fixation (double-sided tape / skin adhesive), remove "
                     "bracelet/watch/rings"),
    ("0:30", "0:40", "Mocap volume check + mask reflections; Calibration take C0 "
                     "(T-pose + per-finger ROM)"),
    ("0:40", "0:45", "cosmik/RGB co-capture check + sync-clap test in both systems"),
    ("0:45", "1:20", "Tasks T0–T5 — one take of 10 reps each "
                     "(instruction + practice + record + save/check ~5-6 min)"),
    ("1:20", "1:30", "Buffer: re-do flagged takes, remove markers, back up files"),
]


def _hdr(ws, row, cols, fill=HDR, font=HDR_FONT):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, WRAP, THIN


def _widths(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def sheet_protocol(wb):
    ws = wb.create_sheet("Protocol")
    ws["A1"] = "Hand teleop — mocap vs fastsam3d: data collection protocol"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "Goal: validate fastsam3d+cosmik markerless hand joint angles against "
        "marker-based mocap (OptiTrack Motive), in ANGLES (frame-independent).",
        f"Subjects: {N_SUBJECTS}  |  Session: ~1h30 each  |  "
        "Collection window: Wed 23 → Fri 25 Jul 2026 (~5 subjects/day).",
        "",
        "MARKER SET (per hand): 21 dorsal markers on the 21 fastsam3d landmarks "
        "— per finger MCP / PIP / DIP / nail (4×5) + wrist. Markers dorsal so a "
        "grip keeps them camera-visible. Place them as close to the joint "
        "creases as possible (just distal to the crease = less soft-tissue slip).",
        "",
        "CO-CAPTURE: record Motive AND cosmik simultaneously for every take. "
        "Start each take with a sharp CLAP (visible in both systems) for sync.",
        "Each task = 2 TAKES of 10 repetitions each (≥20 reps/task; segment "
        "offline). Start every take with the hand still and all markers visible "
        "~1–2 s (seed). The 2 takes give a test–retest / held-out set and guard "
        "against a marker shifting mid-session; the 10 reps give the per-joint "
        "RMSE distribution (mean ± CI) vs SAM3D and the repeatability floor.",
        "",
        "OCCLUSION is the key variable: tasks are graded None → High so we can "
        "plot markerless error vs occlusion. Keep the back of the hand toward "
        "the cameras; avoid pressing the dorsal markers against surfaces.",
        "",
        "MARKER FIXATION: a single marker already fell off in a pilot. Use "
        "double-sided tape + skin adhesive, and CHECK markers between takes "
        "(log 'Marker fell off'). On Wed, pilot the screwdriver/bottle tasks "
        "first to confirm the 21-marker set survives the forceful tasks; if it "
        "doesn't, reduce to a robust subset and note it.",
        "",
        "REFLECTIONS: remove bracelet/watch/rings; set a tight capture volume "
        "and mask static reflections in Motive (a metal bracelet ruined a pilot).",
        "",
        "Per-subject procedure: see the Schedule sheet. Per-take logging: see "
        "the Data log sheet (pre-filled — fill status columns as you go).",
    ]
    for i, t in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=t).alignment = WRAP
    ws.column_dimensions["A"].width = 110
    for i in range(3, 3 + len(lines)):
        ws.row_dimensions[i].height = 30
    return ws


def sheet_tasks(wb):
    ws = wb.create_sheet("Tasks")
    cols = ["ID", "Task", "Description", "Occlusion", "Reps", "Duration", "Why"]
    _hdr(ws, 1, cols)
    for r, t in enumerate(TASKS, 2):
        for j, v in enumerate(t, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = WRAP
            cell.border = THIN
            if j == 1:
                cell.font = BOLD
    _widths(ws, [7, 24, 46, 20, 6, 11, 34])
    ws.freeze_panes = "A2"
    return ws


def sheet_subjects(wb):
    ws = wb.create_sheet("Subjects")
    _hdr(ws, 1, SUBJ_COLS)
    for i in range(1, N_SUBJECTS + 1):
        r = i + 1
        ws.cell(row=r, column=1, value=f"S{i:02d}").font = BOLD
        for j in range(1, len(SUBJ_COLS) + 1):
            ws.cell(row=r, column=j).border = THIN
    hand = DataValidation(type="list", formula1='"Left,Right,Ambidextrous"')
    sex = DataValidation(type="list", formula1='"F,M,Other"')
    yn = DataValidation(type="list", formula1='"Y,N"')
    ws.add_data_validation(hand); ws.add_data_validation(sex)
    ws.add_data_validation(yn)
    hand.add(f"F2:F{N_SUBJECTS+1}"); sex.add(f"H2:H{N_SUBJECTS+1}")
    yn.add(f"K2:K{N_SUBJECTS+1}")
    _widths(ws, [10, 18, 12, 12, 12, 13, 6, 7, 20, 14, 16, 30])
    ws.freeze_panes = "A2"
    return ws


def sheet_log(wb):
    ws = wb.create_sheet("Data log")
    _hdr(ws, 1, LOG_COLS)
    r = 2
    for i in range(1, N_SUBJECTS + 1):
        sid = f"S{i:02d}"
        for t in TASKS:
            tid, reps = t[0], t[4]
            takes = 1 if tid == "C0" else N_TAKES     # 2 takes per task
            for take in range(1, takes + 1):
                ttype = "Calib" if tid == "C0" else "Task"
                row = [sid, ttype, tid, take, "", "", "", reps, "", "", "", "", ""]
                for j, v in enumerate(row, 1):
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.border = THIN
                    if j == 1:
                        cell.font = BOLD
                if tid == "C0":
                    for j in range(1, len(LOG_COLS) + 1):
                        ws.cell(row=r, column=j).fill = SUB
                r += 1
    last = r - 1
    yn = DataValidation(type="list", formula1='"Y,N"')
    qual = DataValidation(type="list", formula1='"Good,Partial,Redo"')
    ws.add_data_validation(yn); ws.add_data_validation(qual)
    for col in ("G", "I", "J", "L"):                 # Y/N columns
        yn.add(f"{col}2:{col}{last}")
    qual.add(f"K2:K{last}")
    _widths(ws, [10, 10, 8, 7, 20, 20, 11, 7, 14, 15, 11, 10, 30])
    ws.freeze_panes = "A2"
    return ws


def sheet_schedule(wb):
    ws = wb.create_sheet("Schedule")
    ws["A1"] = "Per-subject 90-minute schedule"
    ws["A1"].font = Font(bold=True, size=12)
    _hdr(ws, 2, ["From", "To", "Step"])
    for r, (a, b, s) in enumerate(SCHEDULE, 3):
        ws.cell(row=r, column=1, value=a).border = THIN
        ws.cell(row=r, column=2, value=b).border = THIN
        c = ws.cell(row=r, column=3, value=s)
        c.alignment = WRAP; c.border = THIN
    _widths(ws, [8, 8, 90])
    return ws


def sheet_calendar(wb):
    ws = wb.create_sheet("Calendar")
    ws["A1"] = "Subject schedule — 5 subjects/day, Wed 23 → Fri 25 Jul 2026"
    ws["A1"].font = Font(bold=True, size=12)
    _hdr(ws, 2, ["Time slot"] + CAL_DAYS)
    # assign S01..S15 down the days, skipping the lunch row
    sid = 1
    slot_is_lunch = [("lunch" in s) for s in CAL_SLOTS]
    grid = [[None] * len(CAL_DAYS) for _ in CAL_SLOTS]
    for d in range(len(CAL_DAYS)):
        for r in range(len(CAL_SLOTS)):
            if slot_is_lunch[r]:
                grid[r][d] = "— lunch —"
            else:
                grid[r][d] = f"S{sid:02d}"
                sid += 1
    for r, slot in enumerate(CAL_SLOTS):
        row = 3 + r
        c = ws.cell(row=row, column=1, value=slot)
        c.font = BOLD if not slot_is_lunch[r] else Font(italic=True)
        c.border = THIN
        for d in range(len(CAL_DAYS)):
            cell = ws.cell(row=row, column=2 + d, value=grid[r][d])
            cell.alignment = CENTER
            cell.border = THIN
            if slot_is_lunch[r]:
                cell.font = Font(italic=True, color="808080")
            else:
                cell.font = BOLD
                cell.fill = SUB
    _widths(ws, [22, 14, 14, 14])
    ws.cell(row=3 + len(CAL_SLOTS) + 1, column=1,
            value="Each slot = 1h30 (see Schedule tab for the per-subject "
                  "breakdown). Names go in the Subjects tab.").alignment = WRAP
    return ws


def main(out):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_protocol(wb)
    sheet_tasks(wb)
    sheet_calendar(wb)
    sheet_subjects(wb)
    sheet_log(wb)
    sheet_schedule(wb)
    wb.save(out)
    print(f"wrote {out}  ({N_SUBJECTS} subjects, {len(TASKS)} takes/subject, "
          f"{N_SUBJECTS*len(TASKS)} log rows)")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else str(
        Path.home() / "TheophileCodes/MOCAP/hand_data_collection.xlsx")
    main(out)
